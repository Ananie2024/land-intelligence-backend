# app/services/reports/excel_generator.py
"""
Excel Report Generation
Land Intelligence System

Generates Excel reports for tax records, parcels, and dashboard statistics.
"""

import io
import logging
from typing import Any, Dict, List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _apply_header_style(worksheet, row: int) -> None:
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_border(worksheet, min_row: int, max_row: int, min_col: int = 1, max_col: Optional[int] = None) -> None:
    """Apply border to a range of cells."""
    if max_col is None:
        max_col = worksheet.max_column
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            worksheet.cell(row=row, column=col).border = thin_border


def generate_tax_excel(
    parcel_data: Dict[str, Any],
    records: List[Dict[str, Any]],
    title: str = "Tax Assessment Report"
) -> bytes:
    """
    Generate an Excel report for tax data.
    
    Args:
        parcel_data: Parcel information (id, number, owner, etc.)
        records: List of tax assessment records
        title: Report title
        
    Returns:
        Excel file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Report"
    
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="4472C4")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Timestamp
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Parcel Information
    ws.append([])
    ws.append(["Parcel Information"])
    ws["A3"].font = Font(bold=True, color="4472C4")
    
    ws.append(["Parcel Number", parcel_data.get("parcel_number", "N/A")])
    ws.append(["Owner", parcel_data.get("owner_name", "N/A")])
    ws.append(["Area (sqm)", parcel_data.get("area_sqm", 0)])
    ws.append(["Total Outstanding", f"${parcel_data.get('total_outstanding', 0):,.2f}"])
    
    # Tax Records
    ws.append([])
    ws.append(["Assessment Records"])
    ws["A8"].font = Font(bold=True, color="4472C4")
    
    # Header row
    headers = ["Year", "Assessed Value", "Base Tax", "Penalties", "Total", "Status", "Due Date"]
    ws.append(headers)
    _apply_header_style(ws, ws.max_row)
    
    # Data rows
    for record in records:
        ws.append([
            record.get("assessment_year", ""),
            record.get("assessed_value", 0),
            record.get("base_tax_amount", 0),
            record.get("penalties_amount", 0),
            record.get("total_amount", 0),
            record.get("status", ""),
            record.get("due_date", ""),
        ])
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    _apply_border(ws, 4, ws.max_row)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_parcels_excel(
    parcels: List[Dict[str, Any]],
    stats: Optional[Dict[str, Any]] = None,
    title: str = "Land Parcels Report"
) -> bytes:
    """
    Generate an Excel report for parcels data.
    
    Args:
        parcels: List of parcel records
        stats: Optional statistics summary
        title: Report title
        
    Returns:
        Excel file content as bytes
    """
    wb = Workbook()
    
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Title
    summary_ws.merge_cells("A1:E1")
    summary_ws["A1"] = title
    summary_ws["A1"].font = Font(size=16, bold=True, color="4472C4")
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    summary_ws.merge_cells("A2:E2")
    summary_ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    if stats:
        summary_ws.append([])
        summary_ws.append(["Statistics Summary"])
        summary_ws["A3"].font = Font(bold=True, color="4472C4")
        
        summary_ws.append(["Total Parcels", f"{stats.get('total_parcels', 0):,}"])
        summary_ws.append(["Total Area (sqm)", f"{stats.get('total_area_sqm', 0):,.2f}"])
        summary_ws.append(["Total Valuation", f"${stats.get('total_valuation', 0):,.2f}"])
    
    # Parcels sheet
    parcels_ws = wb.create_sheet("Parcels")
    
    headers = ["Parcel #", "Owner", "Area (sqm)", "Valuation", "Parish", "Land Use"]
    parcels_ws.append(headers)
    _apply_header_style(parcels_ws, 1)
    
    for parcel in parcels:
        parcels_ws.append([
            parcel.get("parcel_number", "N/A"),
            parcel.get("owner_name", "N/A"),
            parcel.get("area_sqm", 0),
            parcel.get("valuation", 0) or 0,
            parcel.get("parish_name", "N/A"),
            parcel.get("land_use_category_name", "N/A"),
        ])
    
    # Auto-adjust column widths
    for col in range(1, parcels_ws.max_column + 1):
        parcels_ws.column_dimensions[get_column_letter(col)].width = 18
    
    _apply_border(parcels_ws, 2, parcels_ws.max_row)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_dashboard_excel(
    stats: Dict[str, Any],
    title: str = "System Dashboard Report"
) -> bytes:
    """
    Generate an Excel report for dashboard statistics.
    
    Args:
        stats: Complete system statistics
        title: Report title
        
    Returns:
        Excel file content as bytes
    """
    wb = Workbook()
    
    # Parish Sheet
    parish_ws = wb.active
    parish_ws.title = "Parish Stats"
    
    parish_ws.merge_cells("A1:B1")
    parish_ws["A1"] = "Parish Statistics"
    parish_ws["A1"].font = Font(size=14, bold=True, color="4472C4")
    
    parish_ws.merge_cells("A2:B2")
    parish_ws["A2"] = f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    parish_stats = stats.get("parishes", {})
    parish_ws.append([])
    parish_ws.append(["Total Parishes", parish_stats.get("total_parishes", 0)])
    parish_ws.append(["Total Parcels in Parishes", parish_stats.get("total_parcels", 0)])
    parish_ws.append(["Avg Parcels/Parish", f"{parish_stats.get('avg_parcels_per_parish', 0):,.2f}"])
    
    # Parcel Sheet
    parcel_ws = wb.create_sheet("Parcel Stats")
    
    parcel_ws.merge_cells("A1:B1")
    parcel_ws["A1"] = "Parcel Statistics"
    parcel_ws["A1"].font = Font(size=14, bold=True, color="4472C4")
    
    parcel_stats = stats.get("parcels", {})
    parcel_ws.append([])
    parcel_ws.append(["Total Parcels", parcel_stats.get("total_parcels", 0)])
    parcel_ws.append(["Total Area (sqm)", f"{parcel_stats.get('total_area_sqm', 0):,.2f}"])
    parcel_ws.append(["Total Valuation", f"${parcel_stats.get('total_valuation', 0):,.2f}"])
    parcel_ws.append(["Parcels with Deeds", parcel_stats.get("parcels_with_deeds", 0)])
    
    # User Sheet
    user_ws = wb.create_sheet("User Stats")
    
    user_ws.merge_cells("A1:B1")
    user_ws["A1"] = "User Statistics"
    user_ws["A1"].font = Font(size=14, bold=True, color="4472C4")
    
    user_stats = stats.get("users", {})
    user_ws.append([])
    user_ws.append(["Total Users", user_stats.get("total_users", 0)])
    user_ws.append(["Admin Users", user_stats.get("admin_count", 0)])
    user_ws.append(["Client Users", user_stats.get("client_count", 0)])
    user_ws.append(["Viewer Users", user_stats.get("viewer_count", 0)])
    
    # Document Sheet
    doc_ws = wb.create_sheet("Document Stats")
    
    doc_ws.merge_cells("A1:B1")
    doc_ws["A1"] = "Document Statistics"
    doc_ws["A1"].font = Font(size=14, bold=True, color="4472C4")
    
    doc_stats = stats.get("documents", {})
    doc_ws.append([])
    doc_ws.append(["Total Documents", doc_stats.get("total_documents", 0)])
    doc_ws.append(["Total Size (bytes)", doc_stats.get("total_size_bytes", 0)])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()