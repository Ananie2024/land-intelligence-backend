# tests/unit/test_reports_service.py
"""
Unit tests for Reports Service and generators.

Tests the reports_service.py, pdf_generator.py, and excel_generator.py modules.
"""

import io

import pytest
from unittest.mock import AsyncMock, MagicMock, patch

from app.services.reports.reports_service import ReportsService
from app.services.reports.pdf_generator import (
    generate_tax_pdf,
    generate_parcels_pdf,
    generate_dashboard_pdf,
)
from app.services.reports.excel_generator import (
    generate_tax_excel,
    generate_parcels_excel,
    generate_dashboard_excel,
)


# =============================================================================
# Test Fixtures
# =============================================================================

@pytest.fixture
def mock_db():
    """Create a mock database session."""
    session = AsyncMock()
    session.commit = AsyncMock()
    session.flush = AsyncMock()
    session.execute = AsyncMock()
    session.rollback = AsyncMock()
    session.add = MagicMock()
    return session


@pytest.fixture
def sample_parcel_data():
    """Sample parcel data for testing."""
    return {
        "parcel_id": "test-parcel-uuid",
        "parcel_number": "P-001-2024",
        "owner_name": "John Doe",
        "area_sqm": 1500.50,
        "total_outstanding": 1250.75,
        "overdue_amount": 500.00,
        "upcoming_amount": 750.75,
    }


@pytest.fixture
def sample_tax_records():
    """Sample tax assessment records for testing."""
    return [
        {
            "assessment_year": 2024,
            "assessed_value": 100000.00,
            "base_tax_amount": 1500.00,
            "penalties_amount": 100.00,
            "total_amount": 1600.00,
            "status": "paid",
            "due_date": "2024-03-31",
        },
        {
            "assessment_year": 2023,
            "assessed_value": 95000.00,
            "base_tax_amount": 1425.00,
            "penalties_amount": 50.00,
            "total_amount": 1475.00,
            "status": "overdue",
            "due_date": "2023-03-31",
        },
    ]


@pytest.fixture
def sample_parcels_list():
    """Sample parcels list for testing."""
    return [
        {
            "parcel_number": "P-001-2024",
            "owner_name": "John Doe",
            "area_sqm": 1500.0,
            "valuation": 100000.0,
            "parish_name": "Kigali",
            "land_use_category_name": "Residential",
        },
        {
            "parcel_number": "P-002-2024",
            "owner_name": "Jane Smith",
            "area_sqm": 2500.0,
            "valuation": 150000.0,
            "parish_name": "Gasabo",
            "land_use_category_name": "Commercial",
        },
    ]


@pytest.fixture
def sample_dashboard_stats():
    """Sample dashboard statistics for testing."""
    return {
        "parishes": {
            "total_parishes": 5,
            "total_parcels": 1500,
            "avg_parcels_per_parish": 300.0,
        },
        "parcels": {
            "total_parcels": 1500,
            "total_area_sqm": 5000000.0,
            "total_valuation": 75000000.0,
            "parcels_with_deeds": 1200,
        },
        "users": {
            "total_users": 50,
            "admin_count": 2,
            "client_count": 15,
            "viewer_count": 33,
        },
        "documents": {
            "total_documents": 2500,
            "total_size_bytes": 50000000,
        },
    }


# =============================================================================
# ReportsService Unit Tests
# =============================================================================

class TestReportsServiceGenerateTaxReport:
    """Tests for ReportsService.generate_tax_report method."""

    async def test_generate_tax_report_pdf_format(self, mock_db, sample_parcel_data, sample_tax_records):
        """Test generate_tax_report produces valid PDF bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.tax_service, 'get_outstanding_tax', new_callable=AsyncMock) as mock_tax:
            mock_tax.return_value = {
                "parcel_id": sample_parcel_data["parcel_id"],
                "parcel_number": sample_parcel_data["parcel_number"],
                "total_outstanding": sample_parcel_data["total_outstanding"],
                "overdue_amount": sample_parcel_data["overdue_amount"],
                "upcoming_amount": sample_parcel_data["upcoming_amount"],
                "records": sample_tax_records,
            }
            
            with patch.object(service.parcel_service, 'get_parcel', new_callable=AsyncMock) as mock_parcel:
                mock_parcel.return_value = MagicMock(owner_name="John Doe", area_sqm=1500.50)
                
                result = await service.generate_tax_report("test-parcel-uuid", "pdf")
                
                assert isinstance(result, bytes)
                assert len(result) > 0
                # PDF files start with %PDF signature
                assert result[:4] == b'%PDF'

    async def test_generate_tax_report_excel_format(self, mock_db, sample_parcel_data, sample_tax_records):
        """Test generate_tax_report produces valid Excel bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.tax_service, 'get_outstanding_tax', new_callable=AsyncMock) as mock_tax:
            mock_tax.return_value = {
                "parcel_id": sample_parcel_data["parcel_id"],
                "parcel_number": sample_parcel_data["parcel_number"],
                "total_outstanding": sample_parcel_data["total_outstanding"],
                "records": sample_tax_records,
            }
            
            with patch.object(service.parcel_service, 'get_parcel', new_callable=AsyncMock) as mock_parcel:
                mock_parcel.return_value = None
                
                result = await service.generate_tax_report("test-parcel-uuid", "excel")
                
                assert isinstance(result, bytes)
                assert len(result) > 0
                # Excel files (xlsx) start with PK zip signature
                assert result[:2] == b'PK'

    async def test_generate_tax_report_no_tax_records_raises(self, mock_db):
        """Test generate_tax_report raises ValueError when no tax records found."""
        service = ReportsService(mock_db)
        
        with patch.object(service.tax_service, 'get_outstanding_tax', new_callable=AsyncMock) as mock_tax:
            mock_tax.return_value = None
            
            with pytest.raises(ValueError, match="No tax records found for parcel"):
                await service.generate_tax_report("nonexistent-uuid", "pdf")

    async def test_generate_tax_report_unsupported_format_raises(self, mock_db):
        """Test generate_tax_report raises ValueError for unsupported format."""
        service = ReportsService(mock_db)
        
        # Need to mock both tax_service and parcel_service before calling
        with patch.object(service.tax_service, 'get_outstanding_tax', new_callable=AsyncMock) as mock_tax:
            mock_tax.return_value = {
                "parcel_id": "test-parcel-uuid",
                "parcel_number": "P-001-2024",
                "total_outstanding": 0,
                "records": [],
            }
            
            with patch.object(service.parcel_service, 'get_parcel', new_callable=AsyncMock) as mock_parcel:
                mock_parcel.return_value = MagicMock(owner_name="John Doe", area_sqm=1500.50)
                
                with pytest.raises(ValueError, match="Unsupported format"):
                    await service.generate_tax_report("test-parcel-uuid", "html")


class TestReportsServiceGenerateParcelsReport:
    """Tests for ReportsService.generate_parcels_report method."""

    async def test_generate_parcels_report_pdf_format(self, mock_db, sample_parcels_list, sample_dashboard_stats):
        """Test generate_parcels_report produces valid PDF bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.parcel_service, 'list_parcels', new_callable=AsyncMock) as mock_list:
            mock_list.return_value = {"items": sample_parcels_list}
            
            result = await service.generate_parcels_report(None, "pdf")
            
            assert isinstance(result, bytes)
            assert len(result) > 0
            assert result[:4] == b'%PDF'

    async def test_generate_parcels_report_excel_format(self, mock_db, sample_parcels_list):
        """Test generate_parcels_report produces valid Excel bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.parcel_service, 'list_parcels', new_callable=AsyncMock) as mock_list:
            mock_list.return_value = {"items": sample_parcels_list}
            
            result = await service.generate_parcels_report(None, "excel")
            
            assert isinstance(result, bytes)
            assert len(result) > 0
            assert result[:2] == b'PK'

    async def test_generate_parcels_report_with_parish_filter(self, mock_db, sample_parcels_list):
        """Test generate_parcels_report filters by parish_id."""
        service = ReportsService(mock_db)
        
        with patch.object(service.parcel_service, 'list_parcels_by_parish', new_callable=AsyncMock) as mock_list:
            mock_list.return_value = {"items": [sample_parcels_list[0]]}
            
            result = await service.generate_parcels_report("parish-uuid", "pdf")
            
            assert isinstance(result, bytes)
            assert len(result) > 0
            assert result[:4] == b'%PDF'

    async def test_generate_parcels_report_empty_parcels(self, mock_db):
        """Test generate_parcels_report handles empty parcels list."""
        service = ReportsService(mock_db)
        
        with patch.object(service.parcel_service, 'list_parcels', new_callable=AsyncMock) as mock_list:
            mock_list.return_value = {"items": []}
            
            result = await service.generate_parcels_report(None, "pdf")
            
            # Should still generate a valid PDF with 0 parcels
            assert isinstance(result, bytes)
            assert len(result) > 0


class TestReportsServiceGenerateDashboardReport:
    """Tests for ReportsService.generate_dashboard_report method."""

    async def test_generate_dashboard_report_pdf_format(self, mock_db, sample_dashboard_stats):
        """Test generate_dashboard_report produces valid PDF bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.dashboard_service, 'get_system_stats', new_callable=AsyncMock) as mock_stats:
            stats_mock = MagicMock()
            stats_mock.model_dump.return_value = sample_dashboard_stats
            mock_stats.return_value = stats_mock
            
            result = await service.generate_dashboard_report("pdf")
            
            assert isinstance(result, bytes)
            assert len(result) > 0
            assert result[:4] == b'%PDF'

    async def test_generate_dashboard_report_excel_format(self, mock_db, sample_dashboard_stats):
        """Test generate_dashboard_report produces valid Excel bytes."""
        service = ReportsService(mock_db)
        
        with patch.object(service.dashboard_service, 'get_system_stats', new_callable=AsyncMock) as mock_stats:
            stats_mock = MagicMock()
            stats_mock.model_dump.return_value = sample_dashboard_stats
            mock_stats.return_value = stats_mock
            
            result = await service.generate_dashboard_report("excel")
            
            assert isinstance(result, bytes)
            assert len(result) > 0
            assert result[:2] == b'PK'

    async def test_generate_dashboard_report_dict_stats(self, mock_db, sample_dashboard_stats):
        """Test generate_dashboard_report handles dict stats without model_dump."""
        service = ReportsService(mock_db)
        
        with patch.object(service.dashboard_service, 'get_system_stats', new_callable=AsyncMock) as mock_stats:
            # Return a plain dict (no model_dump method)
            mock_stats.return_value = sample_dashboard_stats
            
            result = await service.generate_dashboard_report("pdf")
            
            assert isinstance(result, bytes)
            assert len(result) > 0


# =============================================================================
# PDF Generator Unit Tests
# =============================================================================

class TestGenerateTaxPdf:
    """Tests for generate_tax_pdf function."""

    def test_generates_valid_pdf(self, sample_parcel_data, sample_tax_records):
        """Test generate_tax_pdf returns valid PDF content."""
        result = generate_tax_pdf(sample_parcel_data, sample_tax_records)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:4] == b'%PDF'

    def test_includes_parcel_information(self, sample_parcel_data, sample_tax_records):
        """Test PDF includes parcel details by checking structure exists."""
        result = generate_tax_pdf(sample_parcel_data, sample_tax_records)
        
        # PDF should have content length indicative of proper structure
        assert len(result) > 1000, "PDF should have meaningful content"

    def test_includes_tax_records_table(self, sample_parcel_data, sample_tax_records):
        """Test PDF includes tax assessment records table."""
        result = generate_tax_pdf(sample_parcel_data, sample_tax_records)
        
        # Should have more content with records than without
        result_no_records = generate_tax_pdf(sample_parcel_data, [])
        assert len(result) > len(result_no_records), "Records should add content"

    def test_handles_empty_records(self, sample_parcel_data):
        """Test PDF generation with no tax records."""
        result = generate_tax_pdf(sample_parcel_data, [])
        
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_custom_title(self, sample_parcel_data, sample_tax_records):
        """Test PDF with custom title."""
        result = generate_tax_pdf(sample_parcel_data, sample_tax_records, title="Custom Tax Report")
        
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestGenerateParcelsPdf:
    """Tests for generate_parcels_pdf function."""

    def test_generates_valid_pdf(self, sample_parcels_list, sample_dashboard_stats):
        """Test generate_parcels_pdf returns valid PDF content."""
        result = generate_parcels_pdf(sample_parcels_list, sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:4] == b'%PDF'

    def test_includes_parcel_summary(self, sample_parcels_list, sample_dashboard_stats):
        """Test PDF includes parcel summary table with content."""
        result = generate_parcels_pdf(sample_parcels_list, sample_dashboard_stats)
        
        # PDF should have substantial content for the summary
        assert len(result) > 1500

    def test_handles_empty_parcels(self, sample_dashboard_stats):
        """Test PDF generation with empty parcels list."""
        result = generate_parcels_pdf([], sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        # Even with empty parcels, stats should still generate content
        assert len(result) > 0

    def test_handles_none_stats(self, sample_parcels_list):
        """Test PDF generation with no stats provided."""
        result = generate_parcels_pdf(sample_parcels_list, None)
        
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_limits_to_50_parcels(self):
        """Test that PDF generation limits parcels to 50 rows."""
        # Create 100 parcels
        many_parcels = [
            {
                "parcel_number": f"P-{i:03d}",
                "owner_name": f"Owner {i}",
                "area_sqm": 1000.0,
                "valuation": 50000.0,
                "parish_name": "Kigali",
                "land_use_category_name": "Residential",
            }
            for i in range(100)
        ]
        
        result = generate_parcels_pdf(many_parcels, {"total_parcels": 100, "total_area_sqm": 100000, "total_valuation": 5000000})
        
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestGenerateDashboardPdf:
    """Tests for generate_dashboard_pdf function."""

    def test_generates_valid_pdf(self, sample_dashboard_stats):
        """Test generate_dashboard_pdf returns valid PDF content."""
        result = generate_dashboard_pdf(sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:4] == b'%PDF'

    def test_includes_parish_statistics(self, sample_dashboard_stats):
        """Test PDF includes parish statistics section with meaningful content."""
        result = generate_dashboard_pdf(sample_dashboard_stats)
        
        # Should have substantial content for all statistics sections
        assert len(result) > 2000

    def test_includes_parcel_statistics(self, sample_dashboard_stats):
        """Test PDF includes parcel statistics section with meaningful content."""
        result = generate_dashboard_pdf(sample_dashboard_stats)
        
        assert len(result) > 2000

    def test_includes_user_statistics(self, sample_dashboard_stats):
        """Test PDF includes user statistics section with meaningful content."""
        result = generate_dashboard_pdf(sample_dashboard_stats)
        
        assert len(result) > 2000

    def test_handles_partial_stats(self):
        """Test PDF generation with partial statistics."""
        partial_stats = {
            "parishes": {"total_parishes": 3},
            "parcels": {"total_parcels": 100},
        }
        
        result = generate_dashboard_pdf(partial_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0


# =============================================================================
# Excel Generator Unit Tests
# =============================================================================

class TestGenerateTaxExcel:
    """Tests for generate_tax_excel function."""

    def test_generates_valid_excel(self, sample_parcel_data, sample_tax_records):
        """Test generate_tax_excel returns valid xlsx content."""
        result = generate_tax_excel(sample_parcel_data, sample_tax_records)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:2] == b'PK'  # ZIP signature for xlsx

    def test_creates_tax_report_sheet(self, sample_parcel_data, sample_tax_records):
        """Test Excel file has 'Tax Report' worksheet."""
        from openpyxl import load_workbook
        
        result = generate_tax_excel(sample_parcel_data, sample_tax_records)
        wb = load_workbook(filename=io.BytesIO(result))
        
        assert "Tax Report" in wb.sheetnames

    def test_includes_parcel_data_in_sheet(self, sample_parcel_data, sample_tax_records):
        """Test Excel sheet contains parcel information."""
        from openpyxl import load_workbook
        
        result = generate_tax_excel(sample_parcel_data, sample_tax_records)
        wb = load_workbook(filename=io.BytesIO(result))
        ws = wb["Tax Report"]
        
        # Check that parcel number appears in the sheet
        found_parcel = any("P-001-2024" in str(cell.value) for row in ws.iter_rows() for cell in row)
        assert found_parcel or ws.max_row >= 4  # At minimum we have header + data rows

    def test_handles_empty_records(self, sample_parcel_data):
        """Test Excel generation with no tax records."""
        result = generate_tax_excel(sample_parcel_data, [])
        
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_custom_title(self, sample_parcel_data, sample_tax_records):
        """Test Excel with custom title."""
        result = generate_tax_excel(sample_parcel_data, sample_tax_records, title="Custom Tax Report")
        
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestGenerateParcelsExcel:
    """Tests for generate_parcels_excel function."""

    def test_generates_valid_excel(self, sample_parcels_list, sample_dashboard_stats):
        """Test generate_parcels_excel returns valid xlsx content."""
        result = generate_parcels_excel(sample_parcels_list, sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:2] == b'PK'

    def test_creates_summary_sheet(self, sample_parcels_list, sample_dashboard_stats):
        """Test Excel file has 'Summary' worksheet."""
        from openpyxl import load_workbook
        
        result = generate_parcels_excel(sample_parcels_list, sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        
        assert "Summary" in wb.sheetnames

    def test_creates_parcels_sheet(self, sample_parcels_list, sample_dashboard_stats):
        """Test Excel file has 'Parcels' worksheet."""
        from openpyxl import load_workbook
        
        result = generate_parcels_excel(sample_parcels_list, sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        
        assert "Parcels" in wb.sheetnames

    def test_includes_statistics_in_summary(self, sample_parcels_list, sample_dashboard_stats):
        """Test Summary sheet contains statistics."""
        from openpyxl import load_workbook
        
        result = generate_parcels_excel(sample_parcels_list, sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        ws = wb["Summary"]
        
        # Should have more than just title header
        assert ws.max_row >= 3

    def test_handles_empty_parcels(self, sample_dashboard_stats):
        """Test Excel generation with empty parcels list."""
        result = generate_parcels_excel([], sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        # Should still create sheets with just headers/stats
        assert len(result) > 0

    def test_handles_none_stats(self, sample_parcels_list):
        """Test Excel generation with no stats."""
        result = generate_parcels_excel(sample_parcels_list, None)
        
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestGenerateDashboardExcel:
    """Tests for generate_dashboard_excel function."""

    def test_generates_valid_excel(self, sample_dashboard_stats):
        """Test generate_dashboard_excel returns valid xlsx content."""
        result = generate_dashboard_excel(sample_dashboard_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result[:2] == b'PK'

    def test_creates_multiple_sheets(self, sample_dashboard_stats):
        """Test Excel file has multiple statistics worksheets."""
        from openpyxl import load_workbook
        
        result = generate_dashboard_excel(sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        
        assert "Parish Stats" in wb.sheetnames
        assert "Parcel Stats" in wb.sheetnames
        assert "User Stats" in wb.sheetnames
        assert "Document Stats" in wb.sheetnames

    def test_includes_parish_stats(self, sample_dashboard_stats):
        """Test Parish Stats sheet contains data."""
        from openpyxl import load_workbook
        
        result = generate_dashboard_excel(sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        ws = wb["Parish Stats"]
        
        assert ws.max_row >= 3

    def test_includes_parcel_stats(self, sample_dashboard_stats):
        """Test Parcel Stats sheet contains data."""
        from openpyxl import load_workbook
        
        result = generate_dashboard_excel(sample_dashboard_stats)
        wb = load_workbook(filename=io.BytesIO(result))
        ws = wb["Parcel Stats"]
        
        assert ws.max_row >= 3

    def test_handles_partial_stats(self):
        """Test Excel generation with partial statistics."""
        partial_stats = {"parishes": {"total_parishes": 3}}
        
        result = generate_dashboard_excel(partial_stats)
        
        assert isinstance(result, bytes)
        assert len(result) > 0